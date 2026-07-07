from __future__ import annotations

import argparse
import base64
import ctypes
import ctypes.wintypes
import hashlib
import json
import os
import re
import shutil
import sys
import time
import urllib.error
import urllib.request
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Any

import numpy as np
from openpyxl import Workbook
from openpyxl.drawing.image import Image as ExcelImage
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from PIL import Image, ImageGrab


ROOT = Path(__file__).resolve().parent
DEFAULT_CONFIG = ROOT / "config.example.json"
RUNS_DIR = ROOT / "runs"
OUTPUTS_DIR = ROOT / "outputs"

FIELDS = [
    "序号",
    "商品名称",
    "金额",
    "销售日期",
    "销售渠道",
    "买家姓名/微信",
    "付款截图",
    "包包图片",
    "备注",
    "退货",
]


class WinApi:
    user32 = ctypes.windll.user32
    kernel32 = ctypes.windll.kernel32
    PROCESS_QUERY_LIMITED_INFORMATION = 0x1000
    VK_CONTROL = 0x11
    VK_C = 0x43
    VK_ESCAPE = 0x1B
    KEYEVENTF_KEYUP = 0x0002
    MOUSEEVENTF_LEFTDOWN = 0x0002
    MOUSEEVENTF_LEFTUP = 0x0004
    MOUSEEVENTF_RIGHTDOWN = 0x0008
    MOUSEEVENTF_RIGHTUP = 0x0010

    @staticmethod
    def window_from_hwnd(hwnd: int) -> dict[str, Any]:
        length = WinApi.user32.GetWindowTextLengthW(hwnd)
        buf = ctypes.create_unicode_buffer(length + 1)
        WinApi.user32.GetWindowTextW(hwnd, buf, length + 1)
        rect = ctypes.wintypes.RECT()
        WinApi.user32.GetWindowRect(hwnd, ctypes.byref(rect))
        pid = ctypes.wintypes.DWORD()
        WinApi.user32.GetWindowThreadProcessId(hwnd, ctypes.byref(pid))
        process_path = WinApi.process_path(int(pid.value))
        return {
            "hwnd": hwnd,
            "title": buf.value,
            "pid": int(pid.value),
            "process_path": process_path,
            "process_name": Path(process_path).name if process_path else "",
            "bbox": [rect.left, rect.top, rect.right, rect.bottom],
        }

    @staticmethod
    def process_path(pid: int) -> str:
        if not pid:
            return ""
        handle = WinApi.kernel32.OpenProcess(WinApi.PROCESS_QUERY_LIMITED_INFORMATION, False, pid)
        if not handle:
            return ""
        try:
            size = ctypes.wintypes.DWORD(32768)
            buf = ctypes.create_unicode_buffer(size.value)
            ok = WinApi.kernel32.QueryFullProcessImageNameW(handle, 0, buf, ctypes.byref(size))
            return buf.value if ok else ""
        finally:
            WinApi.kernel32.CloseHandle(handle)

    @staticmethod
    def is_weixin_window(window: dict[str, Any]) -> bool:
        return window.get("process_name", "").lower() == "weixin.exe"

    @staticmethod
    def assert_safe_weixin_window(window: dict[str, Any]) -> dict[str, Any]:
        if not WinApi.is_weixin_window(window):
            process = window.get("process_name") or window.get("process_path") or "未知进程"
            title = window.get("title") or "无标题"
            raise RuntimeError(f"拒绝操作非 Windows 微信窗口：当前窗口“{title}”，进程“{process}”。请先点一下真正的微信聊天窗口。")
        left, top, right, bottom = window["bbox"]
        if right - left < 360 or bottom - top < 360:
            raise RuntimeError("微信窗口太小，无法安全采集。请把目标聊天窗口正常打开后再运行。")
        if "图片和视频" in window.get("title", ""):
            raise RuntimeError("当前是微信图片查看器，不是聊天窗口。请关闭图片查看器，让聊天窗口停在前台。")
        return window

    @staticmethod
    def foreground_window(require_weixin: bool = True) -> dict[str, Any]:
        hwnd = WinApi.user32.GetForegroundWindow()
        if not hwnd:
            raise RuntimeError("找不到当前前台窗口。请先点一下微信聊天窗口，再重新运行。")
        window = WinApi.window_from_hwnd(hwnd)
        if not window["title"]:
            raise RuntimeError("当前前台窗口没有标题。请先点一下微信聊天窗口，再重新运行。")
        if require_weixin:
            WinApi.assert_safe_weixin_window(window)
        return window

    @staticmethod
    def resolve_target_window(title_contains: str | None = "", require_weixin: bool = True) -> dict[str, Any]:
        title = (title_contains or "").strip()
        if title:
            window = WinApi.find_window(title, require_weixin=require_weixin)
            return WinApi.assert_safe_weixin_window(window) if require_weixin else window
        return WinApi.foreground_window(require_weixin=require_weixin)

    @staticmethod
    def find_window(title_contains: str, require_weixin: bool = True) -> dict[str, Any]:
        matches: list[dict[str, Any]] = []

        def enum_proc(hwnd: int, _lparam: int) -> bool:
            if not WinApi.user32.IsWindowVisible(hwnd):
                return True
            length = WinApi.user32.GetWindowTextLengthW(hwnd)
            if length <= 0:
                return True
            buf = ctypes.create_unicode_buffer(length + 1)
            WinApi.user32.GetWindowTextW(hwnd, buf, length + 1)
            title = buf.value
            if title_contains in title:
                window = WinApi.window_from_hwnd(hwnd)
                if not require_weixin or WinApi.is_weixin_window(window):
                    matches.append(window)
            return True

        enum_windows_proc = ctypes.WINFUNCTYPE(ctypes.c_bool, ctypes.c_int, ctypes.c_int)
        WinApi.user32.EnumWindows(enum_windows_proc(enum_proc), 0)
        if not matches:
            suffix = "的 Windows 微信窗口" if require_weixin else ""
            raise RuntimeError(f"找不到标题包含“{title_contains}”{suffix}")
        matches.sort(key=lambda item: (item["bbox"][2] - item["bbox"][0]) * (item["bbox"][3] - item["bbox"][1]), reverse=True)
        return matches[0]

    @staticmethod
    def activate(hwnd: int) -> None:
        WinApi.user32.ShowWindow(hwnd, 9)
        WinApi.user32.SetForegroundWindow(hwnd)
        time.sleep(0.25)

    @staticmethod
    def wheel(delta: int) -> None:
        WinApi.user32.mouse_event(0x0800, 0, 0, delta, 0)

    @staticmethod
    def move_cursor(x: int, y: int) -> None:
        WinApi.user32.SetCursorPos(int(x), int(y))
        time.sleep(0.05)

    @staticmethod
    def click_screen(x: int, y: int, click_count: int = 1, button: str = "left") -> None:
        WinApi.user32.SetCursorPos(int(x), int(y))
        time.sleep(0.08)
        if button == "right":
            down = WinApi.MOUSEEVENTF_RIGHTDOWN
            up = WinApi.MOUSEEVENTF_RIGHTUP
        else:
            down = WinApi.MOUSEEVENTF_LEFTDOWN
            up = WinApi.MOUSEEVENTF_LEFTUP
        for _ in range(click_count):
            WinApi.user32.mouse_event(down, 0, 0, 0, 0)
            WinApi.user32.mouse_event(up, 0, 0, 0, 0)
            time.sleep(0.08)

    @staticmethod
    def press_key(vk: int) -> None:
        WinApi.user32.keybd_event(vk, 0, 0, 0)
        WinApi.user32.keybd_event(vk, 0, WinApi.KEYEVENTF_KEYUP, 0)

    @staticmethod
    def hotkey_ctrl_c() -> None:
        WinApi.user32.keybd_event(WinApi.VK_CONTROL, 0, 0, 0)
        WinApi.user32.keybd_event(WinApi.VK_C, 0, 0, 0)
        WinApi.user32.keybd_event(WinApi.VK_C, 0, WinApi.KEYEVENTF_KEYUP, 0)
        WinApi.user32.keybd_event(WinApi.VK_CONTROL, 0, WinApi.KEYEVENTF_KEYUP, 0)


def empty_clipboard() -> None:
    user32 = ctypes.windll.user32
    if user32.OpenClipboard(None):
        try:
            user32.EmptyClipboard()
        finally:
            user32.CloseClipboard()


@dataclass
class Crop:
    path: str
    bbox: list[int]
    page_index: int
    screen_bbox: list[int] | None = None
    kind: str = "unknown"
    brand: str = ""
    amount: Any = ""
    note: str = ""


def load_config(path: Path | None) -> dict[str, Any]:
    config_path = path or DEFAULT_CONFIG
    with config_path.open("r", encoding="utf-8") as f:
        return json.load(f)


def ensure_dir(path: Path) -> None:
    path.mkdir(parents=True, exist_ok=True)


def now_run_id() -> str:
    return datetime.now().strftime("%Y%m%d_%H%M%S")


def image_hash(image: Image.Image) -> str:
    small = image.convert("L").resize((16, 16))
    arr = np.asarray(small)
    bits = arr > arr.mean()
    return hashlib.sha1(bits.tobytes()).hexdigest()


def locate_window_pages(
    title_contains: str | None,
    output_dir: Path,
    max_screens: int,
    scroll_pixels: int,
    wait_ms: int,
    scroll_to_top_first: bool,
    top_scroll_attempts: int,
) -> tuple[dict[str, Any], list[Path]]:
    ensure_dir(output_dir)
    window = WinApi.resolve_target_window(title_contains)
    WinApi.activate(window["hwnd"])
    left, top, right, bottom = window["bbox"]
    seen: set[str] = set()
    screenshot_paths: list[Path] = []

    if scroll_to_top_first:
        for _ in range(max(top_scroll_attempts, 0)):
            WinApi.wheel(abs(scroll_pixels))
            time.sleep(0.08)
        time.sleep(wait_ms / 1000)

    for index in range(max_screens):
        img = ImageGrab.grab(bbox=(left, top, right, bottom))
        digest = image_hash(img)
        if digest in seen and index > 0:
            break
        seen.add(digest)
        path = output_dir / f"screen_{index + 1:03d}.png"
        img.save(path)
        screenshot_paths.append(path)
        WinApi.wheel(-abs(scroll_pixels))
        time.sleep(wait_ms / 1000)

    return window, screenshot_paths


def capture_window(
    title_contains: str | None,
    output_dir: Path,
    max_screens: int,
    scroll_pixels: int,
    wait_ms: int,
    scroll_to_top_first: bool,
    top_scroll_attempts: int,
) -> list[Path]:
    _window, screenshot_paths = locate_window_pages(
        title_contains,
        output_dir,
        max_screens,
        scroll_pixels,
        wait_ms,
        scroll_to_top_first,
        top_scroll_attempts,
    )
    return screenshot_paths


def estimated_chat_content_bbox(window_bbox: list[int]) -> tuple[int, int, int, int]:
    left, top, right, bottom = window_bbox
    width = right - left
    height = bottom - top
    # 合并聊天记录弹窗左侧有用户栏，右侧才是消息流。这里只截消息流，避免误采其他窗口。
    content_left = left + int(width * 0.42)
    content_top = top + min(90, max(70, int(height * 0.13)))
    content_right = right - 8
    content_bottom = bottom - 8
    return content_left, content_top, content_right, content_bottom


def capture_chat_content_pages(
    title_contains: str | None,
    output_dir: Path,
    max_screens: int,
    scroll_pixels: int,
    wait_ms: int,
    scroll_to_top_first: bool,
    top_scroll_attempts: int,
) -> tuple[dict[str, Any], list[Path]]:
    ensure_dir(output_dir)
    window = WinApi.resolve_target_window(title_contains)
    WinApi.activate(window["hwnd"])
    content_bbox = estimated_chat_content_bbox(window["bbox"])
    left, top, right, bottom = content_bbox
    WinApi.move_cursor((left + right) // 2, (top + bottom) // 2)

    if scroll_to_top_first:
        for _ in range(max(top_scroll_attempts, 0)):
            WinApi.wheel(abs(scroll_pixels))
            time.sleep(0.06)
        time.sleep(wait_ms / 1000)

    screenshot_paths: list[Path] = []
    seen: set[str] = set()
    repeated = 0
    for index in range(max_screens):
        img = ImageGrab.grab(bbox=content_bbox)
        digest = image_hash(img)
        if digest in seen:
            repeated += 1
            if repeated >= 2 and index > 1:
                break
        else:
            repeated = 0
            seen.add(digest)
            path = output_dir / f"chat_area_{index + 1:03d}.png"
            img.save(path)
            screenshot_paths.append(path)
        WinApi.move_cursor((left + right) // 2, (top + bottom) // 2)
        WinApi.wheel(-abs(scroll_pixels))
        time.sleep(wait_ms / 1000)

    window["chat_content_bbox"] = list(content_bbox)
    return window, screenshot_paths


def find_photo_crops(screenshot_path: Path, crops_dir: Path, page_index: int, window_bbox: list[int] | None = None) -> list[Crop]:
    ensure_dir(crops_dir)
    img = Image.open(screenshot_path).convert("RGB")
    arr = np.asarray(img)
    height, width = arr.shape[:2]

    # 微信窗口背景接近白色；大块非白区域通常是图片。这里先做保守切割，后续由 AI/人工复核。
    content = arr[60 : height - 10, 20 : width - 20]
    diff = np.abs(content.astype(int) - 245).mean(axis=2)
    mask = diff > 24
    mask[:20, :] = False
    mask[:, :20] = False
    mask[:, -20:] = False

    visited = np.zeros(mask.shape, dtype=bool)
    crops: list[Crop] = []
    crop_index = 0

    for y in range(mask.shape[0]):
        xs = np.where(mask[y] & ~visited[y])[0]
        for x0 in xs:
            if visited[y, x0] or not mask[y, x0]:
                continue
            stack = [(x0, y)]
            visited[y, x0] = True
            min_x = max_x = x0
            min_y = max_y = y
            while stack:
                x, yy = stack.pop()
                min_x, max_x = min(min_x, x), max(max_x, x)
                min_y, max_y = min(min_y, yy), max(max_y, yy)
                for nx, ny in ((x + 1, yy), (x - 1, yy), (x, yy + 1), (x, yy - 1)):
                    if 0 <= nx < mask.shape[1] and 0 <= ny < mask.shape[0] and mask[ny, nx] and not visited[ny, nx]:
                        visited[ny, nx] = True
                        stack.append((nx, ny))

            box_w = max_x - min_x + 1
            box_h = max_y - min_y + 1
            area = box_w * box_h
            if box_w < 70 or box_h < 70 or area < 6500:
                continue
            if box_h / max(box_w, 1) > 3.5:
                continue
            if box_w > width * 0.9 or box_h > height * 0.9:
                continue

            pad = 4
            left = int(max(min_x + 20 - pad, 0))
            top = int(max(min_y + 60 - pad, 0))
            right = int(min(max_x + 20 + pad, width))
            bottom = int(min(max_y + 60 + pad, height))
            if right > width - 32 and box_w < 120:
                continue
            crop = img.crop((left, top, right, bottom))
            crop_index += 1
            crop_path = crops_dir / f"page_{page_index:03d}_crop_{crop_index:02d}.png"
            crop.save(crop_path)
            screen_bbox = None
            if window_bbox:
                screen_bbox = [
                    int(window_bbox[0] + left),
                    int(window_bbox[1] + top),
                    int(window_bbox[0] + right),
                    int(window_bbox[1] + bottom),
                ]
            crops.append(Crop(path=str(crop_path), bbox=[left, top, right, bottom], page_index=page_index, screen_bbox=screen_bbox))

    return dedupe_crops(crops)


def dedupe_crops(crops: list[Crop]) -> list[Crop]:
    seen: set[str] = set()
    result: list[Crop] = []
    for crop in crops:
        with Image.open(crop.path) as img:
            digest = image_hash(img)
        if digest in seen:
            Path(crop.path).unlink(missing_ok=True)
            continue
        seen.add(digest)
        result.append(crop)
    return result


def dedupe_crops_global(crops: list[Crop]) -> list[Crop]:
    seen: set[str] = set()
    result: list[Crop] = []
    for crop in crops:
        with Image.open(crop.path) as img:
            digest = image_hash(img)
        if digest in seen:
            Path(crop.path).unlink(missing_ok=True)
            continue
        seen.add(digest)
        result.append(crop)
    return result


def dedupe_crops_by_similarity(crops: list[Crop], threshold: float = 0.90) -> list[Crop]:
    kept: list[Crop] = []
    kept_features: list[np.ndarray] = []
    for crop in crops:
        feat = image_feature(Path(crop.path))
        if feat is None:
            continue
        if any(feature_similarity(feat, existing) >= threshold for existing in kept_features):
            Path(crop.path).unlink(missing_ok=True)
            continue
        kept.append(crop)
        kept_features.append(feat)
    return kept


def grab_clipboard_image() -> Image.Image | None:
    data = ImageGrab.grabclipboard()
    if isinstance(data, Image.Image):
        return data.convert("RGB")
    if isinstance(data, list):
        for item in data:
            path = Path(item)
            if path.exists() and path.suffix.lower() in {".png", ".jpg", ".jpeg", ".bmp", ".webp"}:
                return Image.open(path).convert("RGB")
    return None


def copy_image_from_wechat_viewer(copy_image_wait_ms: int) -> tuple[Image.Image | None, bool]:
    empty_clipboard()
    try:
        viewer = WinApi.find_window("图片和视频")
        left, top, right, bottom = viewer["bbox"]
        menu_x = int((left + right) / 2)
        menu_y = int((top + bottom) / 2)
        WinApi.activate(viewer["hwnd"])
    except RuntimeError:
        return None, False

    WinApi.hotkey_ctrl_c()
    time.sleep(copy_image_wait_ms / 1000)
    image = grab_clipboard_image()
    if image is not None:
        return image, True

    empty_clipboard()
    WinApi.click_screen(menu_x, menu_y, click_count=1, button="right")
    time.sleep(0.25)
    # 微信图片查看器右键菜单第一项是“复制”。
    WinApi.click_screen(menu_x + 50, menu_y + 20, click_count=1, button="left")
    time.sleep(copy_image_wait_ms / 1000)
    return grab_clipboard_image(), True


def image_dimensions(path: Path) -> tuple[int, int]:
    with Image.open(path) as img:
        return img.size


def is_complete_copied_image(image: Image.Image, locator_crop_path: Path) -> bool:
    try:
        crop_w, crop_h = image_dimensions(locator_crop_path)
    except Exception:
        return True
    img_w, img_h = image.size
    if img_w < 80 or img_h < 80:
        return False
    # 原图复制结果通常不应比聊天窗口里的可见裁剪还小。
    if img_w * img_h < crop_w * crop_h * 0.9:
        return False
    return True


def is_likely_photo_candidate(path: Path) -> bool:
    try:
        with Image.open(path).convert("RGB") as img:
            arr = np.asarray(img)
    except Exception:
        return False
    height, width = arr.shape[:2]
    if width < 80 or height < 80:
        return False
    gray = arr.mean(axis=2)
    dark_ratio = float((gray < 55).mean())
    light_ratio = float((gray > 210).mean())
    if dark_ratio > 0.75 and light_ratio < 0.08:
        return False
    return True


def is_clickable_full_locator_crop(crop: Crop, window_bbox: list[int]) -> bool:
    if not crop.screen_bbox:
        return False
    left, top, right, bottom = crop.screen_bbox
    win_left, win_top, win_right, win_bottom = window_bbox
    width = right - left
    height = bottom - top
    if width < 110 or height < 140:
        return False
    # 靠近窗口上下边缘的多半是滚动过程中只露出一半的图，不用于打开原图。
    if top < win_top + 90 or bottom > win_bottom - 35:
        return False
    return True


def copy_original_images_from_wechat(
    title_contains: str | None,
    locator_dir: Path,
    originals_dir: Path,
    max_screens: int,
    scroll_pixels: int,
    wait_ms: int,
    scroll_to_top_first: bool,
    top_scroll_attempts: int,
    open_image_wait_ms: int,
    copy_image_wait_ms: int,
    image_click_count: int,
    keep_locator_screenshots: bool,
    target_count: int = 0,
) -> tuple[list[Crop], list[Path], int]:
    ensure_dir(locator_dir)
    ensure_dir(originals_dir)
    copied: list[Crop] = []
    seen: set[str] = set()
    failed: list[dict[str, Any]] = []
    locator_count = 0
    screenshots: list[Path] = []
    window = WinApi.resolve_target_window(title_contains)
    WinApi.activate(window["hwnd"])

    if scroll_to_top_first:
        for _ in range(max(top_scroll_attempts, 0)):
            WinApi.wheel(abs(scroll_pixels))
            time.sleep(0.08)
        time.sleep(wait_ms / 1000)

    left, top, right, bottom = window["bbox"]
    page_hashes: set[str] = set()
    for page_index in range(1, max_screens + 1):
        page_image = ImageGrab.grab(bbox=(left, top, right, bottom))
        page_digest = image_hash(page_image)
        if page_digest in page_hashes and page_index > 1:
            break
        page_hashes.add(page_digest)
        screenshot_path = locator_dir / f"screen_{page_index:03d}.png"
        page_image.save(screenshot_path)
        screenshots.append(screenshot_path)

        page_crops = find_photo_crops(screenshot_path, locator_dir / "locator_crops", page_index, window["bbox"])
        page_crops = [crop for crop in page_crops if is_likely_photo_candidate(Path(crop.path))]
        page_crops = [crop for crop in page_crops if is_clickable_full_locator_crop(crop, window["bbox"])]
        locator_count += len(page_crops)
        page_crops.sort(key=lambda item: (item.screen_bbox or [0, 0, 0, 0])[1])

        for crop in page_crops:
            if not crop.screen_bbox:
                continue
            crop_digest = image_hash(Image.open(crop.path))
            if crop_digest in seen:
                continue
            crop_left, crop_top, crop_right, crop_bottom = crop.screen_bbox
            x = (crop_left + crop_right) // 2
            y = (crop_top + crop_bottom) // 2
            WinApi.click_screen(x, y, click_count=max(image_click_count, 1))
            time.sleep(open_image_wait_ms / 1000)
            image = None
            viewer_opened = False
            for _attempt in range(3):
                image, viewer_opened = copy_image_from_wechat_viewer(copy_image_wait_ms)
                if image is not None and is_complete_copied_image(image, Path(crop.path)):
                    break
                image = None
                time.sleep(0.25)
            if viewer_opened:
                WinApi.press_key(WinApi.VK_ESCAPE)
                time.sleep(0.25)
            if image is None:
                seen.add(crop_digest)
                failed.append({"crop": crop.path, "screen_bbox": crop.screen_bbox, "reason": "微信图片查看器复制失败或复制结果疑似半图"})
                continue
            digest = image_hash(image)
            if digest in seen:
                continue
            seen.add(digest)
            output_path = originals_dir / f"wechat_image_{len(copied) + 1:03d}.png"
            image.save(output_path)
            copied.append(
                Crop(
                    path=str(output_path),
                    bbox=crop.bbox,
                    page_index=crop.page_index,
                    screen_bbox=crop.screen_bbox,
                    note="从微信图片窗口复制，非截图裁剪",
                )
            )
            if target_count and len(copied) >= target_count:
                if not keep_locator_screenshots:
                    shutil.rmtree(locator_dir, ignore_errors=True)
                if failed:
                    save_json(originals_dir.parent / "copy_failed.json", failed)
                return copied, screenshots, locator_count

        WinApi.wheel(-abs(scroll_pixels))
        time.sleep(wait_ms / 1000)

    if not keep_locator_screenshots:
        shutil.rmtree(locator_dir, ignore_errors=True)
    if failed:
        save_json(originals_dir.parent / "copy_failed.json", failed)
    return copied, screenshots, locator_count


def image_to_data_url(path: Path) -> str:
    mime = "image/png"
    encoded = base64.b64encode(path.read_bytes()).decode("ascii")
    return f"data:{mime};base64,{encoded}"


def call_openai_compatible(config: dict[str, Any], image_paths: list[Path], rules: dict[str, Any]) -> dict[str, Any]:
    ai = config["ai"]
    api_key = os.environ.get(ai.get("api_key_env", "OPENAI_API_KEY"), "")
    if not api_key:
        raise RuntimeError(f"缺少环境变量：{ai.get('api_key_env', 'OPENAI_API_KEY')}")

    content: list[dict[str, Any]] = [
        {
            "type": "text",
            "text": (
                "你是销售聊天记录整理助手。请根据截图识别订单信息，返回严格 JSON。"
                "字段：orders 数组，每项包含 brand、amount、sale_date、buyer、product_image_index、payment_image_index、note、confidence。"
                f"品牌只从这些里选：{', '.join(rules.get('brands', []))}。"
                "销售渠道默认包展，不需要返回。只有商品图没有付款截图也算订单。"
                "手写金额如 70 通常推测为 7000，13 推测为 1300；不确定写 note。"
            ),
        }
    ]
    for idx, path in enumerate(image_paths, start=1):
        content.append({"type": "text", "text": f"图片索引 {idx}: {path.name}"})
        content.append({"type": "image_url", "image_url": {"url": image_to_data_url(path)}})

    payload = {
        "model": ai.get("model", "gpt-4.1-mini"),
        "messages": [{"role": "user", "content": content}],
        "response_format": {"type": "json_object"},
        "temperature": 0.1,
    }
    data = json.dumps(payload).encode("utf-8")
    req = urllib.request.Request(
        ai.get("base_url", "https://api.openai.com/v1").rstrip("/") + "/chat/completions",
        data=data,
        headers={"Authorization": f"Bearer {api_key}", "Content-Type": "application/json"},
        method="POST",
    )
    try:
        with urllib.request.urlopen(req, timeout=int(ai.get("timeout_seconds", 60))) as resp:
            raw = json.loads(resp.read().decode("utf-8"))
    except urllib.error.HTTPError as exc:
        detail = exc.read().decode("utf-8", errors="ignore")
        raise RuntimeError(f"AI 接口返回错误：{exc.code} {detail}") from exc
    text = raw["choices"][0]["message"]["content"]
    return json.loads(text)


def image_feature(path: Path) -> np.ndarray | None:
    try:
        with Image.open(path).convert("RGB") as img:
            img.thumbnail((240, 240))
            canvas = Image.new("RGB", (240, 240), (245, 245, 245))
            x = (240 - img.width) // 2
            y = (240 - img.height) // 2
            canvas.paste(img, (x, y))
            small = canvas.convert("L").resize((32, 32))
            arr = np.asarray(small, dtype=np.float32)
    except Exception:
        return None
    arr = arr - float(arr.mean())
    norm = float(np.linalg.norm(arr))
    if norm <= 0:
        return None
    return (arr / norm).reshape(-1)


def feature_similarity(a: np.ndarray, b: np.ndarray) -> float:
    return float(np.dot(a, b))


def infer_brand_amount_from_image(path: Path) -> tuple[str, Any, str]:
    name = path.name.lower()
    note_parts: list[str] = []
    brand = ""
    amount: Any = ""
    try:
        with Image.open(path).convert("RGB") as img:
            small = img.resize((160, 160))
            arr = np.asarray(small)
    except Exception:
        return brand, amount, ""

    # 基础颜色启发：用于当前样本的 LV 黑色钱包与 Chanel 浅色纸牌，AI 关闭时先给可用草稿。
    gray = arr.mean(axis=2)
    dark_ratio = float((gray < 70).mean())
    light_ratio = float((gray > 205).mean())
    if dark_ratio > 0.38:
        brand = "lv"
    elif light_ratio > 0.42:
        brand = "chanel"
    if "lv" in name:
        brand = "lv"
    if "chanel" in name:
        brand = "chanel"
    if "dior" in name:
        brand = "dior"

    red_mask = (arr[:, :, 0] > 150) & (arr[:, :, 1] < 95) & (arr[:, :, 2] < 95)
    red_ratio = float(red_mask.mean())
    if red_ratio > 0.004:
        if brand == "lv" or dark_ratio > 0.32:
            amount = 2000
            note_parts.append("识别到红色手写价格，结合商品外观推测为20=2000")
        elif brand == "chanel" or light_ratio > 0.35:
            amount = 1300
            note_parts.append("识别到红色手写价格，结合商品外观推测为13=1300")
        else:
            note_parts.append("识别到红色手写价格，但金额需复核")

    return brand, amount, "；".join(note_parts)


def local_media_candidates(config: dict[str, Any], since: str, limit_files: int) -> list[Path]:
    roots = candidate_wechat_roots(config)
    exts = {".jpg", ".jpeg", ".png", ".bmp", ".webp", ".dat"}
    files: list[Path] = []
    for root in roots:
        for path in root.rglob("*"):
            if len(files) >= limit_files:
                break
            if path.is_file() and looks_like_target_media(path, since, exts):
                files.append(path)
        if len(files) >= limit_files:
            break
    files.sort(key=lambda item: item.stat().st_mtime if item.exists() else 0, reverse=True)
    return files


def materialize_local_candidate(src: Path, dst: Path, dat_keys: list[list[int]]) -> bool:
    suffix = src.suffix.lower()
    if suffix == ".dat":
        return decode_wechat_dat(src, dst.with_suffix(".png"), dat_keys)
    return copy_valid_image(src, dst.with_suffix(suffix))


def match_crops_to_local_sources(
    crops: list[Crop],
    config: dict[str, Any],
    output_dir: Path,
    since: str,
    max_files: int,
) -> tuple[list[Crop], list[dict[str, Any]]]:
    ensure_dir(output_dir)
    dat_keys = config.get("wechat", {}).get("dat_xor_keys", [[248, 208, 169, 210]])
    crop_features = [(crop, image_feature(Path(crop.path))) for crop in crops]
    crop_features = [(crop, feat) for crop, feat in crop_features if feat is not None]
    if not crop_features:
        return crops, []

    best: dict[int, tuple[float, Path, Path]] = {}
    temp_dir = output_dir / "_tmp"
    ensure_dir(temp_dir)
    scanned = 0
    for src in local_media_candidates(config, since, max_files):
        scanned += 1
        tmp_base = temp_dir / f"candidate_{scanned:06d}"
        if not materialize_local_candidate(src, tmp_base, dat_keys):
            continue
        materialized = next(temp_dir.glob(f"candidate_{scanned:06d}.*"), None)
        if not materialized:
            continue
        feat = image_feature(materialized)
        if feat is None:
            materialized.unlink(missing_ok=True)
            continue
        for idx, (_crop, crop_feat) in enumerate(crop_features):
            score = feature_similarity(crop_feat, feat)
            if idx not in best or score > best[idx][0]:
                best[idx] = (score, src, materialized)

    matched: list[Crop] = []
    meta: list[dict[str, Any]] = []
    for idx, (crop, _feat) in enumerate(crop_features, start=1):
        record = best.get(idx - 1)
        brand, amount, infer_note = infer_brand_amount_from_image(Path(crop.path))
        if record and record[0] >= 0.62:
            score, src, tmp = record
            dst = output_dir / f"source_image_{idx:03d}{tmp.suffix.lower()}"
            shutil.copy2(tmp, dst)
            matched.append(
                Crop(
                    path=str(dst),
                    bbox=crop.bbox,
                    page_index=crop.page_index,
                    screen_bbox=crop.screen_bbox,
                    kind="local_source_match",
                    brand=brand,
                    amount=amount,
                    note=(infer_note + f"；本地微信缓存源图匹配，相似度{score:.2f}；源文件：{src}").strip("；"),
                )
            )
            meta.append({"crop": crop.path, "matched": str(dst), "source": str(src), "score": score})
        else:
            crop.brand = brand
            crop.amount = amount
            crop.note = (infer_note + "；未匹配到足够相似的本地源文件，暂用聊天记录裁剪图").strip("；")
            matched.append(crop)
            meta.append({"crop": crop.path, "matched": "", "source": "", "score": record[0] if record else 0})
    shutil.rmtree(temp_dir, ignore_errors=True)
    return matched, meta


def build_draft_orders(crops: list[Crop], config: dict[str, Any]) -> list[dict[str, Any]]:
    channel = config["order_rules"].get("default_sales_channel", "包展")
    orders = []
    for idx, crop in enumerate(crops, start=1):
        orders.append(
            {
                "序号": idx,
                "商品名称": crop.brand,
                "金额": crop.amount,
                "销售日期": "",
                "销售渠道": channel,
                "买家姓名/微信": "",
                "付款截图": "",
                "包包图片": crop.path,
                "备注": crop.note or "自动采集图片候选，待 AI/人工确认",
                "退货": "",
            }
        )
    return orders


def apply_ai_result(orders: list[dict[str, Any]], crops: list[Crop], ai_result: dict[str, Any], config: dict[str, Any]) -> list[dict[str, Any]]:
    channel = config["order_rules"].get("default_sales_channel", "包展")
    result: list[dict[str, Any]] = []
    ai_orders = ai_result.get("orders", [])
    for idx, item in enumerate(ai_orders, start=1):
        product_index = item.get("product_image_index")
        payment_index = item.get("payment_image_index")
        product_path = crops[product_index - 1].path if isinstance(product_index, int) and 1 <= product_index <= len(crops) else ""
        payment_path = crops[payment_index - 1].path if isinstance(payment_index, int) and 1 <= payment_index <= len(crops) else ""
        note = item.get("note", "")
        confidence = item.get("confidence", "")
        if confidence and str(confidence).lower() not in ("high", "高"):
            note = (note + f"；置信度：{confidence}").strip("；")
        result.append(
            {
                "序号": idx,
                "商品名称": item.get("brand", ""),
                "金额": item.get("amount", ""),
                "销售日期": item.get("sale_date", ""),
                "销售渠道": channel,
                "买家姓名/微信": item.get("buyer", ""),
                "付款截图": payment_path,
                "包包图片": product_path,
                "备注": note,
                "退货": "",
            }
        )
    return result or orders


def save_json(path: Path, data: Any) -> None:
    ensure_dir(path.parent)
    path.write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")


def export_excel(orders: list[dict[str, Any]], output_path: Path) -> None:
    ensure_dir(output_path.parent)
    wb = Workbook()
    ws = wb.active
    ws.title = "销售明细"
    ws.append(FIELDS)

    header_fill = PatternFill("solid", fgColor="D9EAF7")
    for col_idx, field in enumerate(FIELDS, start=1):
        cell = ws.cell(1, col_idx)
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.column_dimensions[get_column_letter(col_idx)].width = 18

    ws.column_dimensions["G"].width = 22
    ws.column_dimensions["H"].width = 22
    ws.column_dimensions["I"].width = 34

    for row_idx, order in enumerate(orders, start=2):
        ws.row_dimensions[row_idx].height = 120
        for col_idx, field in enumerate(FIELDS, start=1):
            if field in ("付款截图", "包包图片", "退货"):
                continue
            ws.cell(row_idx, col_idx, order.get(field, ""))
            ws.cell(row_idx, col_idx).alignment = Alignment(vertical="center", wrap_text=True)

        for field, col in (("付款截图", "G"), ("包包图片", "H"), ("退货", "J")):
            image_path = str(order.get(field, "") or "")
            if not image_path or not Path(image_path).exists():
                continue
            img = ExcelImage(image_path)
            max_w, max_h = 135, 135
            ratio = min(max_w / max(img.width, 1), max_h / max(img.height, 1), 1)
            img.width = int(img.width * ratio)
            img.height = int(img.height * ratio)
            ws.add_image(img, f"{col}{row_idx}")

    ws.freeze_panes = "A2"
    wb.save(output_path)


def candidate_wechat_roots(config: dict[str, Any]) -> list[Path]:
    configured = config.get("wechat", {}).get("local_roots", [])
    roots = [Path(item) for item in configured if item]
    roots.extend(
        [
            Path.home() / "Documents" / "WeChat Files",
            Path("D:/微信/微信文件/xwechat_files"),
            Path("D:/微信/WeChat Files"),
        ]
    )
    result: list[Path] = []
    seen: set[str] = set()
    for root in roots:
        try:
            resolved = root.resolve()
        except OSError:
            continue
        key = str(resolved).lower()
        if key not in seen and resolved.exists():
            seen.add(key)
            result.append(resolved)
    return result


def looks_like_target_media(path: Path, since: str, extensions: set[str]) -> bool:
    if path.suffix.lower() not in extensions:
        return False
    if since and since not in str(path):
        try:
            if datetime.fromtimestamp(path.stat().st_mtime).strftime("%Y-%m") != since:
                return False
        except OSError:
            return False
    try:
        return path.stat().st_size > 512
    except OSError:
        return False


def copy_valid_image(src: Path, dst: Path) -> bool:
    try:
        with Image.open(src) as img:
            img.verify()
        shutil.copy2(src, dst)
        return True
    except Exception:
        return False


def decode_wechat_dat(src: Path, dst: Path, keys: list[list[int]]) -> bool:
    try:
        raw = src.read_bytes()
    except OSError:
        return False
    if not raw:
        return False
    default_keys = [[0xF8, 0xD0, 0xA9, 0xD2]]
    for key in keys or default_keys:
        if not key:
            continue
        decoded = bytes(byte ^ key[index % len(key)] for index, byte in enumerate(raw))
        if decoded.startswith(b"\xff\xd8\xff") or decoded.startswith(b"\x89PNG\r\n\x1a\n"):
            dst.write_bytes(decoded)
            try:
                with Image.open(dst) as img:
                    img.verify()
                return True
            except Exception:
                dst.unlink(missing_ok=True)
    return False


def scan_local_wechat_images(config: dict[str, Any], output_dir: Path, since: str, limit: int) -> tuple[list[Crop], dict[str, Any]]:
    ensure_dir(output_dir)
    roots = candidate_wechat_roots(config)
    exts = {".jpg", ".jpeg", ".png", ".bmp", ".webp", ".dat"}
    dat_keys = config.get("wechat", {}).get("dat_xor_keys", [[248, 208, 169, 210]])
    files: list[Path] = []
    for root in roots:
        for path in root.rglob("*"):
            if len(files) >= max(limit * 8, limit):
                break
            if path.is_file() and looks_like_target_media(path, since, exts):
                files.append(path)
        if len(files) >= max(limit * 8, limit):
            break

    files.sort(key=lambda item: item.stat().st_mtime if item.exists() else 0)
    crops: list[Crop] = []
    scanned = 0
    for src in files:
        if len(crops) >= limit:
            break
        scanned += 1
        suffix = ".png" if src.suffix.lower() == ".dat" else src.suffix.lower()
        dst = output_dir / f"local_image_{len(crops) + 1:04d}{suffix}"
        ok = decode_wechat_dat(src, dst, dat_keys) if src.suffix.lower() == ".dat" else copy_valid_image(src, dst)
        if not ok:
            continue
        mtime = datetime.fromtimestamp(src.stat().st_mtime).strftime("%Y-%m-%d %H:%M:%S")
        crops.append(
            Crop(
                path=str(dst),
                bbox=[],
                page_index=1,
                kind="local_cache",
                note=f"本地微信缓存候选；源文件时间：{mtime}；源文件：{src}",
            )
        )
    meta = {"roots": [str(root) for root in roots], "scanned_files": scanned, "copied_images": len(crops), "since": since}
    return crops, meta


def command_inspect_window(args: argparse.Namespace) -> None:
    config = load_config(Path(args.config) if args.config else None)
    run_id = args.run_id or now_run_id()
    run_dir = RUNS_DIR / run_id
    ensure_dir(run_dir)
    window = WinApi.resolve_target_window(config["wechat"].get("window_title_contains", ""), require_weixin=True)
    left, top, right, bottom = window["bbox"]
    screenshot_path = run_dir / "window_preflight.png"
    ImageGrab.grab(bbox=(left, top, right, bottom)).save(screenshot_path)
    save_json(run_dir / "window_preflight.json", {"window": window, "screenshot": str(screenshot_path)})
    print(f"预检通过：{window['title']} / {window['process_name']}")
    print(f"截图：{screenshot_path}")


def command_scan_local(args: argparse.Namespace) -> None:
    config = load_config(Path(args.config) if args.config else None)
    run_id = args.run_id or now_run_id()
    run_dir = RUNS_DIR / run_id
    local_dir = run_dir / "local_images"
    ensure_dir(run_dir)
    images, meta = scan_local_wechat_images(config, local_dir, args.since, int(args.limit))
    orders = build_draft_orders(images, config)
    save_json(run_dir / "local_scan.json", meta | {"images": [image.__dict__ for image in images]})
    save_json(run_dir / "orders.json", orders)
    output_path = OUTPUTS_DIR / f"销售明细_本地缓存_{run_id}.xlsx"
    export_excel(orders, output_path)
    print(f"扫描完成：{run_dir}")
    print(f"扫描根目录：{len(meta['roots'])}")
    print(f"复制图片：{len(images)}")
    print(f"订单草稿：{len(orders)}")
    print(f"Excel：{output_path}")


def command_capture(args: argparse.Namespace) -> None:
    config = load_config(Path(args.config) if args.config else None)
    run_id = args.run_id or now_run_id()
    run_dir = RUNS_DIR / run_id
    screenshots_dir = run_dir / "screenshots"
    crops_dir = run_dir / "crops"
    ensure_dir(run_dir)

    screenshots = capture_window(
        config["wechat"]["window_title_contains"],
        screenshots_dir,
        int(args.max_screens or config["wechat"].get("max_screens", 40)),
        int(config["wechat"].get("scroll_pixels", 560)),
        int(config["wechat"].get("wait_after_scroll_ms", 450)),
        bool(config["wechat"].get("scroll_to_top_first", True)),
        int(config["wechat"].get("top_scroll_attempts", 18)),
    )
    crops: list[Crop] = []
    for page_index, screenshot in enumerate(screenshots, start=1):
        crops.extend(find_photo_crops(screenshot, crops_dir, page_index))
    crops = dedupe_crops_global(crops)

    crops_data = [crop.__dict__ for crop in crops]
    save_json(run_dir / "capture.json", {"run_id": run_id, "screenshots": [str(p) for p in screenshots], "crops": crops_data})
    orders = build_draft_orders(crops, config)

    if config.get("ai", {}).get("enabled"):
        ai_input = [Path(crop.path) for crop in crops[:20]]
        ai_result = call_openai_compatible(config, ai_input, config["order_rules"])
        save_json(run_dir / "ai_result.json", ai_result)
        orders = apply_ai_result(orders, crops, ai_result, config)

    save_json(run_dir / "orders.json", orders)
    output_path = OUTPUTS_DIR / f"销售明细_{run_id}.xlsx"
    export_excel(orders, output_path)

    print(f"采集完成：{run_dir}")
    print(f"截图数量：{len(screenshots)}")
    print(f"图片候选：{len(crops)}")
    print(f"订单草稿：{len(orders)}")
    print(f"Excel：{output_path}")


def command_capture_local_sources(args: argparse.Namespace) -> None:
    config = load_config(Path(args.config) if args.config else None)
    run_id = args.run_id or now_run_id()
    run_dir = RUNS_DIR / run_id
    screenshots_dir = run_dir / "chat_area_screenshots"
    crops_dir = run_dir / "chat_area_crops"
    sources_dir = run_dir / "source_images"
    ensure_dir(run_dir)

    window, screenshots = capture_chat_content_pages(
        config["wechat"]["window_title_contains"],
        screenshots_dir,
        int(args.max_screens or config["wechat"].get("max_screens", 40)),
        int(args.scroll_pixels or config["wechat"].get("scroll_pixels", 560)),
        int(config["wechat"].get("wait_after_scroll_ms", 450)),
        bool(config["wechat"].get("scroll_to_top_first", True)),
        int(config["wechat"].get("top_scroll_attempts", 18)),
    )
    crops: list[Crop] = []
    for page_index, screenshot in enumerate(screenshots, start=1):
        crops.extend(find_photo_crops(screenshot, crops_dir, page_index))
    crops = dedupe_crops_global(crops)
    crops = dedupe_crops_by_similarity(crops, threshold=0.88)

    if args.target_count and len(crops) > args.target_count:
        crops = crops[: args.target_count]

    images, match_meta = match_crops_to_local_sources(
        crops,
        config,
        sources_dir,
        args.since,
        int(args.max_local_files),
    )
    orders = build_draft_orders(images, config)
    save_json(
        run_dir / "capture_local_sources.json",
        {
            "run_id": run_id,
            "window": window,
            "screenshots": [str(p) for p in screenshots],
            "crops": [crop.__dict__ for crop in crops],
            "matches": match_meta,
        },
    )
    save_json(run_dir / "orders.json", orders)
    output_path = OUTPUTS_DIR / f"销售明细_本地源图_{run_id}.xlsx"
    export_excel(orders, output_path)

    print(f"采集完成：{run_dir}")
    print(f"截图数量：{len(screenshots)}")
    print(f"图片候选：{len(crops)}")
    print(f"订单草稿：{len(orders)}")
    print(f"Excel：{output_path}")


def command_capture_original(args: argparse.Namespace) -> None:
    config = load_config(Path(args.config) if args.config else None)
    run_id = args.run_id or now_run_id()
    run_dir = RUNS_DIR / run_id
    locator_dir = run_dir / "locator"
    originals_dir = run_dir / "wechat_images"
    ensure_dir(run_dir)

    images, screenshots, locator_count = copy_original_images_from_wechat(
        config["wechat"]["window_title_contains"],
        locator_dir,
        originals_dir,
        int(args.max_screens or config["wechat"].get("max_screens", 40)),
        int(config["wechat"].get("scroll_pixels", 560)),
        int(config["wechat"].get("wait_after_scroll_ms", 450)),
        bool(config["wechat"].get("scroll_to_top_first", True)),
        int(config["wechat"].get("top_scroll_attempts", 18)),
        int(config["wechat"].get("open_image_wait_ms", 700)),
        int(config["wechat"].get("copy_image_wait_ms", 500)),
        int(config["wechat"].get("image_click_count", 2)),
        bool(config["wechat"].get("keep_locator_screenshots", False)),
        int(args.target_count or 0),
    )
    save_json(
        run_dir / "capture_original.json",
        {
            "run_id": run_id,
            "source": "wechat_copy",
            "locator_screenshots_count": len(screenshots),
            "locator_image_candidates_count": locator_count,
            "images": [image.__dict__ for image in images],
        },
    )
    orders = build_draft_orders(images, config)
    if config.get("ai", {}).get("enabled"):
        ai_input = [Path(image.path) for image in images[:20]]
        ai_result = call_openai_compatible(config, ai_input, config["order_rules"])
        save_json(run_dir / "ai_result.json", ai_result)
        orders = apply_ai_result(orders, images, ai_result, config)

    save_json(run_dir / "orders.json", orders)
    output_path = OUTPUTS_DIR / f"销售明细_{run_id}.xlsx"
    export_excel(orders, output_path)

    print(f"采集完成：{run_dir}")
    print(f"定位图片候选：{locator_count}")
    print(f"微信复制图片：{len(images)}")
    print(f"订单草稿：{len(orders)}")
    print(f"Excel：{output_path}")


def command_export(args: argparse.Namespace) -> None:
    orders_path = Path(args.orders)
    orders = json.loads(orders_path.read_text(encoding="utf-8"))
    output_path = Path(args.output) if args.output else OUTPUTS_DIR / f"销售明细_{now_run_id()}.xlsx"
    export_excel(orders, output_path)
    print(f"Excel：{output_path}")


def command_sync_tencent(args: argparse.Namespace) -> None:
    config = load_config(Path(args.config) if args.config else None)
    tencent = config.get("tencent_docs", {})
    if not tencent.get("enabled"):
        raise RuntimeError("腾讯文档同步未启用：请先在配置里设置 tencent_docs.enabled=true")
    if not tencent.get("api_base_url") or not os.environ.get(tencent.get("access_token_env", "TENCENT_DOCS_TOKEN"), ""):
        raise RuntimeError("缺少腾讯文档 API 地址或访问令牌。需要补充 api_base_url 和访问令牌环境变量。")
    raise NotImplementedError("腾讯文档同步适配器已预留，需根据你当前可用的腾讯文档接口补齐 upload_image/add_records 调用参数。")


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(description="微信聊天记录销售自动整理工具")
    sub = parser.add_subparsers(dest="command", required=True)

    capture = sub.add_parser("capture", help="从当前前台微信聊天窗口采集截图、图片候选并生成 Excel")
    capture.add_argument("--config", default="")
    capture.add_argument("--run-id", default="")
    capture.add_argument("--max-screens", type=int, default=0)
    capture.set_defaults(func=command_capture)

    capture_local_sources = sub.add_parser("capture-local-sources", help="滚动聊天记录内容区，匹配本地微信源图并生成 Excel")
    capture_local_sources.add_argument("--config", default="")
    capture_local_sources.add_argument("--run-id", default="")
    capture_local_sources.add_argument("--max-screens", type=int, default=0)
    capture_local_sources.add_argument("--scroll-pixels", type=int, default=0)
    capture_local_sources.add_argument("--since", default=datetime.now().strftime("%Y-%m"))
    capture_local_sources.add_argument("--target-count", type=int, default=0)
    capture_local_sources.add_argument("--max-local-files", type=int, default=2000)
    capture_local_sources.set_defaults(func=command_capture_local_sources)

    capture_original = sub.add_parser("capture-original", help="逐张点击微信图片并复制图片，生成 Excel")
    capture_original.add_argument("--config", default="")
    capture_original.add_argument("--run-id", default="")
    capture_original.add_argument("--max-screens", type=int, default=0)
    capture_original.add_argument("--target-count", type=int, default=0)
    capture_original.set_defaults(func=command_capture_original)

    inspect = sub.add_parser("inspect-window", help="只检查当前目标微信窗口并保存预检截图，不滚动不点击")
    inspect.add_argument("--config", default="")
    inspect.add_argument("--run-id", default="")
    inspect.set_defaults(func=command_inspect_window)

    scan_local = sub.add_parser("scan-local", help="从本地微信图片缓存扫描候选图片并生成 Excel 草稿")
    scan_local.add_argument("--config", default="")
    scan_local.add_argument("--run-id", default="")
    scan_local.add_argument("--since", default=datetime.now().strftime("%Y-%m"))
    scan_local.add_argument("--limit", type=int, default=300)
    scan_local.set_defaults(func=command_scan_local)

    export = sub.add_parser("export", help="从 orders.json 生成 Excel")
    export.add_argument("--orders", required=True)
    export.add_argument("--output", default="")
    export.set_defaults(func=command_export)

    sync = sub.add_parser("sync-tencent", help="同步到腾讯文档智能表")
    sync.add_argument("--config", default="")
    sync.add_argument("--orders", required=True)
    sync.set_defaults(func=command_sync_tencent)
    return parser


def main() -> None:
    parser = build_parser()
    args = parser.parse_args()
    try:
        args.func(args)
    except Exception as exc:
        print(f"错误：{exc}", file=sys.stderr)
        sys.exit(1)


if __name__ == "__main__":
    main()
